import subprocess
import sys
import os
import time
import random
import json
from datetime import datetime

import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", message=".*dict.*")
warnings.filterwarnings("ignore", message=".*Pydantic.*")

def install_dependencies():
    packages = {
        "pandas": "pandas",
        "jobspy": "python-jobspy",
        "openpyxl": "openpyxl",
        "bs4": "beautifulsoup4",
        "duckduckgo_search": "duckduckgo-search",
        "tqdm": "tqdm"
    }
    missing = []
    for module_name, pip_name in packages.items():
        try:
            __import__(module_name)
        except ImportError:
            missing.append(pip_name)

    if missing:
        print(f"Installing missing dependencies: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing, "--quiet"])

install_dependencies()

import pandas as pd
from jobspy import scrape_jobs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from hn_scraper import fetch_hn_jobs
from twitter_scraper import fetch_twitter_jobs
from remote_boards_scraper import fetch_remote_boards
from ats_scraper import fetch_ats_jobs
from true_remote_filter import filter_true_remote
from ai_ml_filter import filter_ai_ml_jobs
from state_manager import filter_new_jobs

with open("config.json", "r") as f:
    config = json.load(f)

SEARCH_TERMS = config.get("search_terms", [])
TECH_SKILLS = config.get("tech_skills", [])
PLATFORMS = config.get("jobspy_platforms", ["linkedin", "indeed"])
TARGET_LOCATIONS = config.get("jobspy_target_locations", ["USA", "Europe", "UK"])
RESULTS_WANTED = config.get("jobspy_results_wanted", 150)
HOURS_OLD = config.get("jobspy_hours_old", 720)

OUTPUT_DIR = os.environ.get("AI_JOBS_OUTPUT_DIR", ".")

def main():
    all_data = []
    total_raw_results = 0

    print("=== AI/ML Remote Jobs Dataset Scraper ===")
    print(f"Started at: {datetime.now().isoformat()}")

    # PHASE 1: Community Sources
    try:
        hn_df = fetch_hn_jobs(TECH_SKILLS)
        if hn_df is not None and not hn_df.empty:
            all_data.append(hn_df)
            total_raw_results += len(hn_df)
            print(f"Scraped {len(hn_df)} HackerNews jobs")
    except Exception as e:
        print(f"Warning: Failed to scrape HackerNews. Error: {e}")

    try:
        tw_df = fetch_twitter_jobs(TECH_SKILLS)
        if tw_df is not None and not tw_df.empty:
            all_data.append(tw_df)
            total_raw_results += len(tw_df)
            print(f"Scraped {len(tw_df)} X/Twitter jobs")
    except Exception as e:
        print(f"Warning: Failed to scrape X/Twitter. Error: {e}")

    # PHASE 2: Remote Niche Job Boards
    try:
        boards_df = fetch_remote_boards(TECH_SKILLS)
        if boards_df is not None and not boards_df.empty:
            all_data.append(boards_df)
            total_raw_results += len(boards_df)
            print(f"Scraped {len(boards_df)} remote board jobs")
    except Exception as e:
        print(f"Warning: Failed to scrape Remote Niche Boards. Error: {e}")

    # PHASE 3: ATS Direct Platforms
    try:
        ats_df = fetch_ats_jobs(TECH_SKILLS)
        if ats_df is not None and not ats_df.empty:
            all_data.append(ats_df)
            total_raw_results += len(ats_df)
            print(f"Scraped {len(ats_df)} ATS platform jobs")
    except Exception as e:
        print(f"Warning: Failed to scrape ATS Platforms. Error: {e}")

    # PHASE 4: Mainstream Boards (LinkedIn, Indeed, Glassdoor)
    for loc in TARGET_LOCATIONS:
        for term in SEARCH_TERMS:
            for platform in PLATFORMS:
                print(f"Scraping {platform} for '{term}' in '{loc}'...")
                try:
                    kwargs = {
                        "site_name": [platform],
                        "search_term": term,
                        "location": loc,
                        "results_wanted": RESULTS_WANTED,
                        "is_remote": True,
                        "hours_old": HOURS_OLD
                    }
                    if platform == "indeed" and loc.lower() in ["usa", "uk", "canada"]:
                        kwargs["country_indeed"] = loc.lower()

                    jobs = scrape_jobs(**kwargs)
                    count = len(jobs) if jobs is not None and not jobs.empty else 0
                    print(f"Scraped {count} results from {platform}")

                    if count > 0:
                        all_data.append(jobs)
                        total_raw_results += count
                except Exception as e:
                    print(f"Warning: Failed to scrape {platform} for '{term}' in '{loc}'. Error: {e}")

                time.sleep(random.uniform(3, 7))

    print(f"\nTotal raw results collected: {total_raw_results}")

    if not all_data:
        print("No data collected across any platform. Exiting.")
        return

    # Merge all datasets
    df = pd.concat(all_data, ignore_index=True)

    # Standardize columns
    COLUMNS_TO_KEEP = {
        'company': 'Company Name',
        'title': 'Job Title',
        'location': 'Location',
        'job_url': 'Job URL',
        'date_posted': 'Date Posted',
        'site': 'Site',
        'company_industry': 'Company Industry',
        'company_num_employees': 'Company Employee Count',
        'min_amount': 'Salary Min',
        'max_amount': 'Salary Max',
        'description': 'Description'
    }
    for col in COLUMNS_TO_KEEP.keys():
        if col not in df.columns:
            df[col] = None

    df = df[list(COLUMNS_TO_KEEP.keys())]
    df.rename(columns=COLUMNS_TO_KEEP, inplace=True)

    # Clean null companies
    df['Company Name'] = df['Company Name'].fillna("").astype(str)
    df = df[df['Company Name'].str.strip() != ""]
    df = df[~df['Company Name'].str.strip().str.lower().isin(["nan", "none"])]

    # Deduplicate
    df = df.drop_duplicates(subset=['Company Name', 'Job Title'], keep='first')
    df = df.reset_index(drop=True)
    print(f"After deduplication: {len(df)} unique jobs")

    # True Remote Filter
    df = filter_true_remote(df)

    # AI/ML Relevance Filter
    df, ai_stats = filter_ai_ml_jobs(df)
    print(f"AI/ML Filter: {ai_stats['removed']} non-AI jobs removed, {ai_stats['total_after_filter']} AI/ML jobs remaining")

    # State Manager - remove previously seen
    original_size = len(df)
    df = filter_new_jobs(df)
    print(f"State Manager: Blocked {original_size - len(df)} duplicate/previously-seen jobs.")

    print(f"Final valid NEW AI/ML jobs to save: {len(df)}")

    if df.empty:
        print("No AI/ML jobs found this run. Exiting.")
        return

    # Generate output files
    date_str = datetime.now().strftime("%Y-%m-%d_%H%M")

    # CSV output (primary)
    csv_filename = os.path.join(OUTPUT_DIR, f"ai-ml-remote-jobs_{date_str}.csv")
    try:
        df.to_csv(csv_filename, index=False)
        print(f"CSV saved: {csv_filename}")
    except Exception as e:
        print(f"Warning: Failed to save CSV. Error: {e}")

    # Latest copy
    latest_csv = os.path.join(OUTPUT_DIR, "ai-ml-remote-jobs_latest.csv")
    try:
        df.to_csv(latest_csv, index=False)
        print(f"Latest CSV saved: {latest_csv}")
    except Exception as e:
        print(f"Warning: Failed to save latest CSV. Error: {e}")

    # JSON output
    json_filename = os.path.join(OUTPUT_DIR, f"ai-ml-remote-jobs_{date_str}.json")
    try:
        df.to_json(json_filename, orient="records", date_format="iso")
        print(f"JSON saved: {json_filename}")
    except Exception as e:
        print(f"Warning: Failed to save JSON. Error: {e}")

    # Excel output (formatted)
    xlsx_filename = os.path.join(OUTPUT_DIR, f"ai-ml-remote-jobs_{date_str}.xlsx")
    try:
        if 'Date Posted' in df.columns:
            try:
                df['Date Posted'] = pd.to_datetime(df['Date Posted']).dt.tz_localize(None)
            except:
                pass

        with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="AI ML Remote Jobs")

        wb = openpyxl.load_workbook(xlsx_filename)
        ws = wb["AI ML Remote Jobs"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                val_str = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(val_str))
            ws.column_dimensions[column].width = min(max_length + 2, 80)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(xlsx_filename)
        print(f"Excel saved: {xlsx_filename}")
    except Exception as e:
        print(f"Warning: Failed to save Excel. Error: {e}")

    # Summary stats
    print(f"\n=== Dataset Summary ===")
    print(f"Date: {date_str}")
    print(f"Total AI/ML jobs: {len(df)}")
    print(f"Files: {csv_filename}, {json_filename}, {xlsx_filename}")
    print(f"Completed at: {datetime.now().isoformat()}")

if __name__ == "__main__":
    main()
